"""Export metadata records into Excel workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from metadata_app.services.xml_service import XmlService
from metadata_app.models.metadata_record import MetadataRecord
from metadata_app.config import SECTION_COLORS


class ExportService:
    """Handles conversion of metadata records to Excel spreadsheets."""

    def __init__(self, xml_service: XmlService) -> None:
        self._xml_service = xml_service

    def export_folder(self, folder: Path, destination: Path) -> Path:
        """Export all XML files in the given folder to a single Excel workbook."""
        records = self._collect_records(Path(folder))
        if not records:
            raise FileNotFoundError(f"No XML files found in {folder}")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metadata"

        header = ["#", "Media Type", "Title", "Section", "Field", "Value"]
        sheet.append(header)
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row_index = 2
        sequence_number = 1

        for _, record in records:
            for section in record.sections:
                section_color_hex = (section.color or SECTION_COLORS.get(section.name, "")).replace("#", "")
                fill = None
                if section_color_hex:
                    if len(section_color_hex) == 6:
                        section_color_hex = f"FF{section_color_hex.upper()}"
                    fill = PatternFill(start_color=section_color_hex, end_color=section_color_hex, fill_type="solid")
                for field_name, value in section.fields.items():
                    sheet.append(
                        [
                            sequence_number,
                            record.media_type,
                            record.title,
                            section.name,
                            field_name,
                            value,
                        ]
                    )
                    if fill:
                        sheet.cell(row=row_index, column=4).fill = fill
                    row_index += 1
                    sequence_number += 1

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max(12, min(max_length + 2, 60))
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        output_path = self._normalize_destination(destination)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _collect_records(self, folder: Path) -> List[Tuple[Path, MetadataRecord]]:
        """Return XML records found in the folder."""
        if not folder.exists():
            return []
        paths = sorted(folder.glob("*.xml"))
        records: List[Tuple[Path, MetadataRecord]] = []
        for path in paths:
            try:
                records.append((path, self._xml_service.load_record(path)))
            except Exception:
                continue
        return records

    def _normalize_destination(self, destination: Path) -> Path:
        """Ensure the export destination is a file path."""
        destination = Path(destination)
        if destination.is_dir():
            return destination / "metadata_export.xlsx"
        if destination.suffix.lower() != ".xlsx":
            return destination.with_suffix(".xlsx")
        return destination
